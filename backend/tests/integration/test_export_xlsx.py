"""Verify Excel export sheet structure matches the production audit contract."""

from __future__ import annotations

import re
from datetime import UTC, datetime, timedelta
from io import BytesIO
from uuid import UUID

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.live_room import LiveRoom
from app.models.response import Response
from app.models.session_question import SessionQuestion
from app.services.response_service import ResponseService
from app.services.results_service import (
    IST,
    ResultsService,
    _ms_between,
    format_export_timestamp,
)
from tests.integration.test_scoring import _close_and_reveal, _setup_open_question

EVERY_ANSWERS_HEADERS = (
    "Participant Name",
    "Participant ID",
    "Room ID",
    "Quiz ID",
    "Question Number",
    "Question ID",
    "Question Text",
    "Selected Option",
    "Correct Option",
    "Correct",
    "Question Broadcast Timestamp",
    "Question Open Timestamp",
    "Answer Submitted Timestamp",
    "Answer Order",
    "Response Time (milliseconds)",
    "Base Score",
    "Time Bonus",
    "Streak Bonus",
    "Total Awarded",
    "Streak Before",
    "Streak After",
    "Rank Before Submission",
    "Rank After Submission",
)

TIMELINE_HEADERS = (
    "Timestamp",
    "Event",
    "Participant",
    "Question Number",
    "Question ID",
    "Selected Option",
    "Details",
)

IST_TIMESTAMP_RE = re.compile(
    r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}\.\d{3} IST$"
)


def _parse_export_ts(value: str) -> datetime:
    """Parse ``YYYY-MM-DD HH:mm:ss.SSS IST`` back to an aware datetime."""
    assert IST_TIMESTAMP_RE.match(value), value
    assert not value.endswith("Z")
    body = value.removesuffix(" IST")
    dt = datetime.strptime(body, "%Y-%m-%d %H:%M:%S.%f")
    return dt.replace(tzinfo=IST)


def _assert_ist_timestamp(value: object) -> str:
    assert isinstance(value, str), value
    assert not value.endswith("Z"), value
    assert "+00:00" not in value, value
    assert IST_TIMESTAMP_RE.match(value), value
    return value


def test_export_xlsx_sheets_and_columns(
    client: TestClient,
    admin_token: str,
    db_session: Session,
) -> None:
    room, option_ids, [joined] = _setup_open_question(
        client,
        admin_token,
        db_session,
        title="Export XLSX",
        joiners=[("Ann", "ann@example.com")],
    )
    room_id = UUID(room["id"])
    participant_id = UUID(joined["participant"]["id"])
    ResponseService(db_session).submit(
        room_id=room_id,
        participant_id=participant_id,
        option_ids=[UUID(option_ids[0])],
        require_connected=False,
    )
    _close_and_reveal(db_session, room_id)

    raw = ResultsService(db_session).export_xlsx(room_id)
    wb = load_workbook(BytesIO(raw))
    assert wb.sheetnames == ["Participants", "Every Answers", "Timeline"]

    answers = list(wb["Every Answers"].iter_rows(values_only=True))
    assert answers[0] == EVERY_ANSWERS_HEADERS
    row = answers[1]
    live_room = db_session.get(LiveRoom, room_id)
    assert live_room is not None

    assert row[0] == "Ann"
    assert row[1] == str(participant_id)
    assert row[2] == str(room_id)
    assert row[3] == str(live_room.quiz_id)
    assert row[4] == 1
    assert row[9] is True
    _assert_ist_timestamp(row[10])
    _assert_ist_timestamp(row[11])
    _assert_ist_timestamp(row[12])
    assert row[13] == 1
    assert row[15] == 10
    assert row[18] == 10
    assert row[19] == 0
    assert row[20] == 1
    assert row[21] == 1
    assert row[22] == 1

    db_response = (
        db_session.query(Response).filter(Response.participant_id == participant_id).one()
    )
    question = db_session.get(SessionQuestion, db_response.session_question_id)
    assert question is not None
    assert db_response.submitted_at is not None
    assert row[12] == format_export_timestamp(db_response.submitted_at)

    # UTC → IST is +05:30
    utc_submitted = db_response.submitted_at
    if utc_submitted.tzinfo is None:
        utc_submitted = utc_submitted.replace(tzinfo=UTC)
    else:
        utc_submitted = utc_submitted.astimezone(UTC)
    ist_expected = utc_submitted + timedelta(hours=5, minutes=30)
    parsed = _parse_export_ts(str(row[12]))
    assert parsed.hour == ist_expected.hour or abs(
        (parsed.astimezone(UTC) - utc_submitted).total_seconds()
    ) < 1

    broadcast = question.broadcast_at or question.opened_at
    assert broadcast is not None
    expected_ms = _ms_between(broadcast, db_response.submitted_at)
    assert row[14] == expected_ms
    assert abs(int(db_response.response_time_ms or 0) - expected_ms) <= 1

    timeline = list(wb["Timeline"].iter_rows(values_only=True))
    assert timeline[0] == TIMELINE_HEADERS
    for event_row in timeline[1:]:
        if event_row[0]:
            _assert_ist_timestamp(event_row[0])
    events = {r[1] for r in timeline[1:]}
    assert "Answer Submitted" in events
    assert "Question Broadcast" in events
    assert "Question Open" in events
    answer_rows = [r for r in timeline[1:] if r[1] == "Answer Submitted"]
    assert answer_rows[0][2] == "Ann"
    assert answer_rows[0][3] == 1
    assert answer_rows[0][4]
    assert answer_rows[0][5]


def test_format_export_timestamp_is_ist_not_utc() -> None:
    dt = datetime(2026, 8, 4, 12, 0, 0, 123456, tzinfo=UTC)
    formatted = format_export_timestamp(dt)
    assert formatted == "2026-08-04 17:30:00.123 IST"
    assert formatted is not None
    assert not formatted.endswith("Z")
    assert "+00:00" not in formatted


def test_export_answer_order_and_response_time_for_multiple_participants(
    client: TestClient,
    admin_token: str,
    db_session: Session,
) -> None:
    room, option_ids, joiners = _setup_open_question(
        client,
        admin_token,
        db_session,
        title="Export Order",
        joiners=[
            ("Alice", "alice@example.com"),
            ("Bob", "bob@example.com"),
            ("Charlie", "charlie@example.com"),
        ],
    )
    room_id = UUID(room["id"])

    for joiner, opt_idx in (
        (joiners[0], 0),
        (joiners[1], 1),
        (joiners[2], 0),
    ):
        ResponseService(db_session).submit(
            room_id=room_id,
            participant_id=UUID(joiner["participant"]["id"]),
            option_ids=[UUID(option_ids[opt_idx])],
            require_connected=False,
        )

    _close_and_reveal(db_session, room_id)

    wb = load_workbook(BytesIO(ResultsService(db_session).export_xlsx(room_id)))
    answers = list(wb["Every Answers"].iter_rows(values_only=True))
    by_name = {row[0]: row for row in answers[1:]}

    assert by_name["Alice"][13] == 1
    assert by_name["Bob"][13] == 2
    assert by_name["Charlie"][13] == 3

    ordered = sorted(answers[1:], key=lambda r: int(r[13]))
    times = [_parse_export_ts(str(r[12])) for r in ordered]
    assert times[0] <= times[1] <= times[2]

    for row in answers[1:]:
        _assert_ist_timestamp(row[10])
        _assert_ist_timestamp(row[12])
        broadcast = _parse_export_ts(str(row[10]))
        submitted = _parse_export_ts(str(row[12]))
        expected = max(0, int((submitted - broadcast).total_seconds() * 1000))
        assert row[14] == expected

    timeline = list(wb["Timeline"].iter_rows(values_only=True))
    submitted_events = [r for r in timeline[1:] if r[1] == "Answer Submitted"]
    assert len(submitted_events) == 3
    assert {r[2] for r in submitted_events} == {"Alice", "Bob", "Charlie"}
    for event_row in submitted_events:
        _assert_ist_timestamp(event_row[0])


def test_export_xlsx_ranks_and_streaks(
    client: TestClient,
    admin_token: str,
    db_session: Session,
) -> None:
    room, option_ids, joiners = _setup_open_question(
        client,
        admin_token,
        db_session,
        title="Export Ranks",
        joiners=[("Ann", "ann@example.com"), ("Bob", "bob@example.com")],
    )
    room_id = UUID(room["id"])
    ResponseService(db_session).submit(
        room_id=room_id,
        participant_id=UUID(joiners[0]["participant"]["id"]),
        option_ids=[UUID(option_ids[0])],
        require_connected=False,
    )
    ResponseService(db_session).submit(
        room_id=room_id,
        participant_id=UUID(joiners[1]["participant"]["id"]),
        option_ids=[UUID(option_ids[1])],
        require_connected=False,
    )
    _close_and_reveal(db_session, room_id)

    wb = load_workbook(BytesIO(ResultsService(db_session).export_xlsx(room_id)))
    answers = list(wb["Every Answers"].iter_rows(values_only=True))
    by_name = {row[0]: row for row in answers[1:]}
    assert by_name["Ann"][9] is True
    assert by_name["Bob"][9] is False
    assert by_name["Ann"][19] == 0
    assert by_name["Ann"][20] == 1
    assert by_name["Bob"][19] == 0
    assert by_name["Bob"][20] == 0
    assert by_name["Ann"][21] == 1
    assert by_name["Ann"][22] == 1
    assert by_name["Bob"][22] == 2

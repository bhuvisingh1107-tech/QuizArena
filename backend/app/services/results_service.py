"""Read-only session results and analytics (admin)."""

from __future__ import annotations

import csv
import io
from collections import defaultdict
from dataclasses import dataclass
from datetime import UTC, datetime
from typing import Any
from uuid import UUID

from sqlalchemy.orm import Session, selectinload
from sqlalchemy import select

from app.core.exceptions import NotFoundError
from app.models.participant import Participant
from app.models.response import Response
from app.models.session_question import SessionQuestion
from app.repositories.live_room_repository import LiveRoomRepository
from app.repositories.participant_repository import ParticipantRepository
from app.schemas.results import (
    LeaderboardEntryData,
    OptionDistributionData,
    PodiumData,
    QuestionAnalyticsData,
    ResultsData,
    ResultsRoomData,
    ResultsSummaryData,
    SectionAnalyticsData,
)


@dataclass(frozen=True)
class RankedParticipant:
    participant: Participant
    rank: int


def format_export_timestamp(value: datetime | None) -> str | None:
    """ISO-8601 UTC with millisecond precision for Excel audit columns."""
    if value is None:
        return None
    dt = value
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=UTC)
    else:
        dt = dt.astimezone(UTC)
    return dt.isoformat(timespec="milliseconds").replace("+00:00", "Z")


def _as_uuid(value: Any) -> UUID | None:
    if value is None:
        return None
    try:
        return UUID(str(value))
    except (TypeError, ValueError):
        return None


def assign_competition_ranks(participants: list[Participant]) -> list[RankedParticipant]:
    """Sort by score DESC, joined_at ASC; assign competition ranks by score."""
    ordered = sorted(
        participants,
        key=lambda p: (-int(p.total_score or 0), p.joined_at),
    )
    ranked: list[RankedParticipant] = []
    i = 0
    while i < len(ordered):
        score = int(ordered[i].total_score or 0)
        j = i + 1
        while j < len(ordered) and int(ordered[j].total_score or 0) == score:
            j += 1
        competition_rank = i + 1
        for k in range(i, j):
            ranked.append(RankedParticipant(participant=ordered[k], rank=competition_rank))
        i = j
    return ranked


class ResultsService:
    """Compute room results, analytics, and CSV export from persisted data."""

    def __init__(self, session: Session) -> None:
        self._session = session
        self._rooms = LiveRoomRepository(session)
        self._participants = ParticipantRepository(session)

    def get_results(self, room_id: UUID) -> ResultsData:
        room = self._rooms.get_by_id(room_id)
        if room is None:
            raise NotFoundError("NOT_FOUND", "Live room not found")

        participants = self._participants.list_for_room(room_id)
        questions = sorted(room.session_questions, key=lambda q: q.sort_order)
        sections = sorted(room.session_sections, key=lambda s: s.sort_order)
        section_name_by_id = {s.id: s.name for s in sections}

        responses = self._list_responses_for_room(room_id)
        responses_by_question: dict[UUID, list[Response]] = defaultdict(list)
        responses_by_participant: dict[UUID, list[Response]] = defaultdict(list)
        for response in responses:
            responses_by_question[response.session_question_id].append(response)
            responses_by_participant[response.participant_id].append(response)

        ranked = assign_competition_ranks(participants)
        leaderboard = [self._leaderboard_entry(rp) for rp in ranked]
        podium = PodiumData(entries=leaderboard[:3])

        total_questions = len(questions)
        participant_count = len(participants)
        average_score = (
            sum(int(p.total_score or 0) for p in participants) / participant_count
            if participant_count
            else 0.0
        )
        average_accuracy = self._average_accuracy(participants, total_questions)
        average_response_time = self._average_response_time(responses)

        question_analytics = [
            self._question_analytics(
                question=question,
                question_index=index,
                section_name=section_name_by_id.get(question.session_section_id, ""),
                question_responses=responses_by_question.get(question.id, []),
                participant_count=participant_count,
            )
            for index, question in enumerate(questions)
        ]

        section_analytics = [
            self._section_analytics(
                section_id=section.id,
                name=section.name,
                section_questions=[q for q in questions if q.session_section_id == section.id],
                participants=participants,
                responses_by_participant=responses_by_participant,
            )
            for section in sections
        ]

        return ResultsData(
            room=ResultsRoomData(
                id=room.id,
                room_code=room.room_code,
                quiz_title_snapshot=room.quiz_title_snapshot,
                state=room.state,
                started_at=room.started_at,
                completed_at=room.completed_at,
            ),
            summary=ResultsSummaryData(
                participant_count=participant_count,
                average_score=round(average_score, 2),
                average_accuracy_percent=round(average_accuracy, 2),
                total_questions=total_questions,
                average_response_time_ms=average_response_time,
            ),
            leaderboard=leaderboard,
            podium=podium,
            question_analytics=question_analytics,
            section_analytics=section_analytics,
        )

    def list_participants_admin(self, room_id: UUID) -> tuple[list[Participant], list[int], int]:
        """Return participants with competition ranks for admin listing."""
        room = self._rooms.get_by_id(room_id)
        if room is None:
            raise NotFoundError("NOT_FOUND", "Live room not found")
        participants = self._participants.list_for_room(room_id)
        ranked = assign_competition_ranks(participants)
        return (
            [rp.participant for rp in ranked],
            [rp.rank for rp in ranked],
            len(ranked),
        )

    def export_csv(self, room_id: UUID) -> str:
        """Build CSV text for leaderboard export."""
        room = self._rooms.get_by_id(room_id)
        if room is None:
            raise NotFoundError("NOT_FOUND", "Live room not found")

        participants = self._participants.list_for_room(room_id)
        ranked = assign_competition_ranks(participants)

        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(
            [
                "Rank",
                "Display Name",
                "Email",
                "Score",
                "Correct",
                "Incorrect",
                "Unanswered",
                "Streak",
            ]
        )
        for rp in ranked:
            p = rp.participant
            writer.writerow(
                [
                    rp.rank,
                    p.display_name,
                    p.email,
                    int(p.total_score or 0),
                    int(p.total_correct or 0),
                    int(p.total_incorrect or 0),
                    int(p.unanswered_count or 0),
                    int(p.streak or 0),
                ]
            )
        return buffer.getvalue()

    def export_xlsx(self, room_id: UUID) -> bytes:
        """Build multi-sheet Excel workbook for session results export."""
        from openpyxl import Workbook

        room = self._rooms.get_by_id(room_id)
        if room is None:
            raise NotFoundError("NOT_FOUND", "Live room not found")

        participants = self._participants.list_for_room(room_id)
        ranked = assign_competition_ranks(participants)
        questions = sorted(room.session_questions, key=lambda q: q.sort_order)
        responses = self._list_responses_for_room(room_id)
        responses_by_participant: dict[UUID, list[Response]] = defaultdict(list)
        for response in responses:
            responses_by_participant[response.participant_id].append(response)

        question_index_by_id = {q.id: idx for idx, q in enumerate(questions)}
        options_by_question: dict[UUID, dict[UUID, str]] = {}
        correct_options_by_question: dict[UUID, list[str]] = {}
        for question in questions:
            opts = {opt.id: opt.text for opt in question.options}
            options_by_question[question.id] = opts
            correct_options_by_question[question.id] = [
                opt.text
                for opt in sorted(question.options, key=lambda o: o.sort_order)
                if opt.is_correct
            ]

        wb = Workbook()

        # ── Sheet 1: Participants ─────────────────────────────────────────
        ws_p = wb.active
        ws_p.title = "Participants"
        ws_p.append(
            [
                "Participant",
                "Rank",
                "Score",
                "Correct",
                "Incorrect",
                "Accuracy",
                "Average Response Time",
                "Fastest Response",
                "Longest Streak",
            ]
        )
        total_q = len(questions)
        for rp in ranked:
            p = rp.participant
            ordered = sorted(
                responses_by_participant.get(p.id, []),
                key=lambda r: question_index_by_id.get(r.session_question_id, 0),
            )
            answered = [
                r
                for r in ordered
                if not r.is_unanswered and r.submitted_at is not None
            ]
            times = [
                int(r.response_time_ms)
                for r in answered
                if r.response_time_ms is not None
            ]
            avg_time = round(sum(times) / len(times), 2) if times else None
            fastest = min(times) if times else None
            correct = int(p.total_correct or 0)
            incorrect = int(p.total_incorrect or 0)
            attempted = correct + incorrect
            accuracy = round((correct / attempted) * 100.0, 2) if attempted else 0.0
            if total_q and attempted == 0:
                accuracy = 0.0
            ws_p.append(
                [
                    p.display_name,
                    rp.rank,
                    int(p.total_score or 0),
                    correct,
                    incorrect,
                    accuracy,
                    avg_time,
                    fastest,
                    self._longest_streak(ordered),
                ]
            )

        # ── Sheet 2: Every Answers ────────────────────────────────────────
        ws_q = wb.create_sheet("Every Answers")
        ws_q.append(
            [
                "Participant Name",
                "Participant ID",
                "Question Number",
                "Question ID",
                "Question Text",
                "Selected Option",
                "Correct Option",
                "Correct",
                "Question Shown Timestamp",
                "Answer Submitted Timestamp",
                "Response Time (milliseconds)",
                "Base Score",
                "Time Bonus",
                "Streak Bonus",
                "Total Score Awarded",
                "Rank Before Submission",
                "Rank After Submission",
            ]
        )
        participant_name = {p.id: p.display_name for p in participants}
        for response in sorted(
            responses,
            key=lambda r: (
                participant_name.get(r.participant_id, ""),
                question_index_by_id.get(r.session_question_id, 0),
            ),
        ):
            qid = response.session_question_id
            q_index = question_index_by_id.get(qid)
            question = next((q for q in questions if q.id == qid), None)
            opt_map = options_by_question.get(qid, {})
            selected_texts: list[str] = []
            for raw_id in response.selected_option_ids or []:
                try:
                    selected_texts.append(opt_map.get(UUID(str(raw_id)), str(raw_id)))
                except (TypeError, ValueError):
                    selected_texts.append(str(raw_id))
            if response.is_unanswered or response.submitted_at is None:
                correct_cell = False
            else:
                correct_cell = bool(response.is_correct)
            shown_at = question.opened_at if question is not None else None
            ws_q.append(
                [
                    participant_name.get(response.participant_id, ""),
                    str(response.participant_id),
                    (q_index + 1) if q_index is not None else None,
                    str(qid),
                    question.prompt_text if question is not None else "",
                    "; ".join(selected_texts),
                    "; ".join(correct_options_by_question.get(qid, [])),
                    correct_cell,
                    format_export_timestamp(shown_at),
                    format_export_timestamp(response.submitted_at),
                    response.response_time_ms,
                    int(response.base_points_earned or 0),
                    int(response.time_bonus_earned or 0),
                    int(response.streak_bonus_earned or 0),
                    int(response.total_points_earned or 0),
                    response.rank_before,
                    response.rank_after,
                ]
            )

        # ── Sheet 3: Timeline ─────────────────────────────────────────────
        ws_t = wb.create_sheet("Timeline")
        ws_t.append(
            [
                "Timestamp",
                "Event",
                "Participant",
                "Question Number",
                "Selected Option",
            ]
        )
        for row in self._timeline_rows(room, responses, questions, participants):
            ws_t.append(row)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def _longest_streak(ordered_responses: list[Response]) -> int:
        streak = 0
        best = 0
        for response in ordered_responses:
            if response.is_unanswered:
                streak = 0
            elif response.is_correct:
                streak += 1
                best = max(best, streak)
            else:
                streak = 0
        return best

    def _timeline_rows(
        self,
        room,
        responses: list[Response],
        questions: list[SessionQuestion],
        participants: list[Participant] | None = None,
    ) -> list[list[Any]]:
        """Prefer session_events; otherwise synthesize from room + responses.

        Columns: Timestamp | Event | Participant | Question Number | Selected Option
        """
        from app.models.session_event import SessionEvent
        from app.services.session_event_service import ANSWER_SUBMITTED, TIMELINE_LABELS

        name_by_id = {p.id: p.display_name for p in (participants or [])}
        question_number_by_id = {q.id: q.sort_order + 1 for q in questions}
        options_by_question: dict[UUID, dict[UUID, str]] = {
            q.id: {opt.id: opt.text for opt in q.options} for q in questions
        }

        events = list(
            self._session.scalars(
                select(SessionEvent)
                .where(SessionEvent.live_room_id == room.id)
                .order_by(SessionEvent.created_at.asc())
            ).all()
        )
        if events:
            rows: list[list[Any]] = []
            for event in events:
                label = TIMELINE_LABELS.get(
                    event.event_type,
                    event.event_type.replace("_", " ").title(),
                )
                payload = event.payload_json if isinstance(event.payload_json, dict) else {}
                participant = (
                    payload.get("displayName")
                    or name_by_id.get(_as_uuid(payload.get("participantId")), "")
                    or ""
                )
                question_number = payload.get("questionNumber")
                if question_number is None and payload.get("questionIndex") is not None:
                    try:
                        question_number = int(payload["questionIndex"]) + 1
                    except (TypeError, ValueError):
                        question_number = None
                selected_option = payload.get("selectedOption") or ""
                ts = payload.get("submittedAt") if event.event_type == ANSWER_SUBMITTED else None
                if not ts:
                    ts = format_export_timestamp(event.created_at)
                rows.append(
                    [
                        ts,
                        label,
                        participant,
                        question_number,
                        selected_option,
                    ]
                )
            return rows

        synthetic: list[tuple[Any, str, str, Any, str]] = []
        if room.created_at:
            synthetic.append((room.created_at, "Room Created", "", None, ""))
        for participant in participants or []:
            if participant.joined_at:
                synthetic.append(
                    (
                        participant.joined_at,
                        "Participant Joined",
                        participant.display_name,
                        None,
                        "",
                    )
                )
        if room.started_at:
            synthetic.append((room.started_at, "Quiz Started", "", None, ""))
        for question in questions:
            if question.opened_at:
                synthetic.append(
                    (
                        question.opened_at,
                        "Question Shown",
                        "",
                        question.sort_order + 1,
                        "",
                    )
                )
        for response in responses:
            if response.submitted_at is None:
                continue
            qid = response.session_question_id
            opt_map = options_by_question.get(qid, {})
            selected_texts: list[str] = []
            for raw_id in response.selected_option_ids or []:
                try:
                    selected_texts.append(opt_map.get(UUID(str(raw_id)), str(raw_id)))
                except (TypeError, ValueError):
                    selected_texts.append(str(raw_id))
            synthetic.append(
                (
                    response.submitted_at,
                    "Answer Submitted",
                    name_by_id.get(response.participant_id, str(response.participant_id)),
                    question_number_by_id.get(qid),
                    "; ".join(selected_texts),
                )
            )
        if room.completed_at:
            synthetic.append((room.completed_at, "Quiz Ended", "", None, ""))

        synthetic.sort(key=lambda item: item[0] or datetime.min.replace(tzinfo=UTC))
        return [
            [
                format_export_timestamp(ts) if hasattr(ts, "isoformat") else ts,
                label,
                participant,
                question_number,
                selected_option,
            ]
            for ts, label, participant, question_number, selected_option in synthetic
        ]

    def _list_responses_for_room(self, room_id: UUID) -> list[Response]:
        stmt = (
            select(Response)
            .join(Participant, Response.participant_id == Participant.id)
            .where(Participant.live_room_id == room_id)
            .options(selectinload(Response.session_question))
        )
        return list(self._session.scalars(stmt).all())

    @staticmethod
    def _leaderboard_entry(rp: RankedParticipant) -> LeaderboardEntryData:
        p = rp.participant
        return LeaderboardEntryData(
            rank=rp.rank,
            participant_id=p.id,
            display_name=p.display_name,
            score=int(p.total_score or 0),
            streak=int(p.streak or 0),
            total_correct=int(p.total_correct or 0),
            total_incorrect=int(p.total_incorrect or 0),
            unanswered_count=int(p.unanswered_count or 0),
        )

    @staticmethod
    def _average_accuracy(participants: list[Participant], total_questions: int) -> float:
        if not participants or total_questions <= 0:
            return 0.0
        accuracies: list[float] = []
        for p in participants:
            correct = int(p.total_correct or 0)
            accuracies.append((correct / total_questions) * 100.0)
        return sum(accuracies) / len(accuracies)

    @staticmethod
    def _average_response_time(responses: list[Response]) -> float | None:
        times = [
            int(r.response_time_ms)
            for r in responses
            if r.response_time_ms is not None and not r.is_unanswered
        ]
        if not times:
            return None
        return round(sum(times) / len(times), 2)

    @staticmethod
    def _question_analytics(
        *,
        question: SessionQuestion,
        question_index: int,
        section_name: str,
        question_responses: list[Response],
        participant_count: int,
    ) -> QuestionAnalyticsData:
        correct = 0
        incorrect = 0
        answered = 0
        times: list[int] = []
        selection_counts: dict[UUID, int] = defaultdict(int)

        for response in question_responses:
            if response.is_unanswered or response.submitted_at is None:
                continue
            answered += 1
            if response.is_correct:
                correct += 1
            else:
                incorrect += 1
            if response.response_time_ms is not None:
                times.append(int(response.response_time_ms))
            for raw_id in response.selected_option_ids or []:
                try:
                    selection_counts[UUID(str(raw_id))] += 1
                except (TypeError, ValueError):
                    continue

        unanswered = max(0, participant_count - answered)

        answered_for_accuracy = correct + incorrect
        accuracy = (correct / answered_for_accuracy * 100.0) if answered_for_accuracy else 0.0
        avg_time = round(sum(times) / len(times), 2) if times else None

        options = sorted(question.options, key=lambda o: o.sort_order)
        distribution = [
            OptionDistributionData(
                option_id=opt.id,
                text=opt.text,
                selected_count=selection_counts.get(opt.id, 0),
                is_correct=bool(opt.is_correct),
            )
            for opt in options
        ]

        return QuestionAnalyticsData(
            question_id=question.id,
            question_index=question_index,
            prompt_text=question.prompt_text,
            section_name=section_name,
            submission_count=answered,
            correct_count=correct,
            incorrect_count=incorrect,
            unanswered_count=unanswered,
            accuracy_percent=round(accuracy, 2),
            average_response_time_ms=avg_time,
            option_distribution=distribution,
        )

    @staticmethod
    def _section_analytics(
        *,
        section_id: UUID,
        name: str,
        section_questions: list[SessionQuestion],
        participants: list[Participant],
        responses_by_participant: dict[UUID, list[Response]],
    ) -> SectionAnalyticsData:
        question_ids = {q.id for q in section_questions}
        question_count = len(section_questions)
        if not participants or not question_ids:
            return SectionAnalyticsData(
                section_id=section_id,
                name=name,
                average_score=0.0,
                question_count=question_count,
            )

        section_scores: list[float] = []
        for participant in participants:
            total = 0
            for response in responses_by_participant.get(participant.id, []):
                if response.session_question_id in question_ids:
                    total += int(response.total_points_earned or 0)
            section_scores.append(float(total))

        average = sum(section_scores) / len(section_scores)
        return SectionAnalyticsData(
            section_id=section_id,
            name=name,
            average_score=round(average, 2),
            question_count=question_count,
        )
